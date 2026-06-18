#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from PyQt6.QtWidgets import *
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor

import json

from core.database import db
from core.auth import auth_manager
from config import SAMPLE_STATUS


class DeafGeneSampleAlreadyExistsException(Exception):
    pass


class DeafGeneSampleValidationException(Exception):
    pass


class DeafGeneSampleManagement(QWidget):
    def __init__(self):
        super().__init__()
        self._deafGeneCurFilter = {}
        self._deafGeneLastSearchTime = None
        self.setupDeafGeneSampleUI()
        self.refreshDeafGeneSampleList()
        
    def setupDeafGeneSampleUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title_label = QLabel("耳聋基因样本管理")
        title_label.setStyleSheet("QLabel { color: #333; font-size: 18px; font-weight: bold; }")
        layout.addWidget(title_label)
        
        toolbar = self.buildDeafGeneActionToolbar()
        layout.addWidget(toolbar)
        
        search_area = self.buildDeafGeneSearchBox()
        layout.addWidget(search_area)
        
        self.deaf_gene_sample_table = self.buildDeafGeneSampleGrid()
        layout.addWidget(self.deaf_gene_sample_table)
        
    def buildDeafGeneActionToolbar(self):
        toolbar = QFrame()
        toolbar.setStyleSheet("background-color: white; border-radius: 8px; padding: 10px;")
        
        layout = QHBoxLayout(toolbar)
        layout.setContentsMargins(10, 5, 10, 5)
        
        import_btn = QPushButton("📥 批量导入")
        import_btn.clicked.connect(self.importDeafGeneExcelData)
        layout.addWidget(import_btn)
        
        add_btn = QPushButton("➕ 新增样本")
        add_btn.clicked.connect(self.openDeafGeneAddDialog)
        layout.addWidget(add_btn)
        
        layout.addStretch()
        
        export_btn = QPushButton("📤 导出数据")
        export_btn.clicked.connect(self.exportDeafGeneToExcel)
        layout.addWidget(export_btn)
        
        return toolbar
        
    def buildDeafGeneSearchBox(self):
        search_frame = QFrame()
        search_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 10px;")
        
        layout = QGridLayout(search_frame)
        layout.setContentsMargins(10, 5, 10, 5)
        layout.setSpacing(10)
        
        layout.addWidget(QLabel("样本编号:"), 0, 0)
        self.deaf_gene_sample_no_input = QLineEdit()
        self.deaf_gene_sample_no_input.setPlaceholderText("请输入样本编号")
        self.deaf_gene_sample_no_input.setFixedWidth(200)
        layout.addWidget(self.deaf_gene_sample_no_input, 0, 1)
        
        layout.addWidget(QLabel("受检者姓名:"), 0, 2)
        self.deaf_gene_patient_name_input = QLineEdit()
        self.deaf_gene_patient_name_input.setPlaceholderText("请输入姓名")
        self.deaf_gene_patient_name_input.setFixedWidth(150)
        layout.addWidget(self.deaf_gene_patient_name_input, 0, 3)
        
        layout.addWidget(QLabel("状态:"), 0, 4)
        self.deaf_gene_status_filter = QComboBox()
        self.deaf_gene_status_filter.addItem("全部", None)
        for status_key, status_name in SAMPLE_STATUS.items():
            self.deaf_gene_status_filter.addItem(status_name, status_key)
        self.deaf_gene_status_filter.setFixedWidth(120)
        layout.addWidget(self.deaf_gene_status_filter, 0, 5)
        
        search_btn = QPushButton("🔍 搜索")
        search_btn.clicked.connect(self.runDeafGeneSearch)
        layout.addWidget(search_btn, 0, 6)
        
        reset_btn = QPushButton("🔄 重置")
        reset_btn.clicked.connect(self.clearDeafGeneSearchInputs)
        layout.addWidget(reset_btn, 0, 7)
        
        return search_frame
        
    def buildDeafGeneSampleGrid(self):
        table = QTableWidget()
        table.setStyleSheet("""
            QTableWidget { background-color: white; border-radius: 8px; gridline-color: #eee; color: #333; }
            QTableWidget::item { padding: 8px; color: #333; }
            QTableWidget::item:selected { background-color: #0078d4; color: white; }
        """)
        
        headers = [
            "样本编号", "受检者姓名", "性别", "年龄", 
            "临床诊断", "送检单位", "检测项目", "状态", "操作"
        ]
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        
        table.verticalHeader().setDefaultSectionSize(40)
        table.verticalHeader().setVisible(False)
        
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        
        table.cellDoubleClicked.connect(self.openDeafGeneEditDialog)
        
        return table
        
    def refreshDeafGeneSampleList(self, filters=None):
        samples = db.get_samples(filters)
        self.fillDeafGeneSampleRows(samples)
    
    def fillDeafGeneSampleRows(self, samples):
        self.deaf_gene_sample_table.setRowCount(len(samples))
        
        for row, sample in enumerate(samples):
            self._fillDeafGeneSampleRow(row, sample)
            self._addDeafGeneActionButtons(row, sample)
    
    def _fillDeafGeneSampleRow(self, row, sample):
        self.deaf_gene_sample_table.setItem(row, 0, QTableWidgetItem(sample['sample_no']))
        self.deaf_gene_sample_table.setItem(row, 1, QTableWidgetItem(sample['patient_name']))
        self.deaf_gene_sample_table.setItem(row, 2, QTableWidgetItem(sample.get('gender', '')))
        self.deaf_gene_sample_table.setItem(row, 3, QTableWidgetItem(sample.get('age', '')))
        self.deaf_gene_sample_table.setItem(row, 4, QTableWidgetItem(sample.get('clinical_diagnosis', '')))
        self.deaf_gene_sample_table.setItem(row, 5, QTableWidgetItem(sample.get('hospital', '')))
        self.deaf_gene_sample_table.setItem(row, 6, QTableWidgetItem(sample.get('test_project', '')))
        
        status_item = QTableWidgetItem(SAMPLE_STATUS.get(sample['status'], sample['status']))
        self._colorDeafGeneStatusText(status_item, sample['status'])
        self.deaf_gene_sample_table.setItem(row, 7, status_item)
    
    def _colorDeafGeneStatusText(self, item, status):
        color_map = {
            'pending': QColor('#ff9800'),
            'analyzing': QColor('#2196f3'),
            'completed': QColor('#4caf50'),
            'failed': QColor('#f44336')
        }
        
        color = color_map.get(status, QColor('#666'))
        item.setForeground(color)
    
    def _addDeafGeneActionButtons(self, row, sample):
        button_widget = QWidget()
        layout = QHBoxLayout(button_widget)
        layout.setContentsMargins(5, 2, 5, 2)
        layout.setSpacing(5)
        
        edit_btn = self._buildDeafGeneEditBtn(row)
        layout.addWidget(edit_btn)
        
        delete_btn = self._buildDeafGeneDeleteBtn(sample['id'])
        layout.addWidget(delete_btn)
        
        self.deaf_gene_sample_table.setCellWidget(row, 8, button_widget)
    
    def _buildDeafGeneEditBtn(self, row):
        edit_btn = QPushButton("编辑")
        edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #ffffff;
                color: #333333;
                border: 1px solid #cccccc;
                border-radius: 3px;
                padding: 4px 8px;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #f0f0f0; border-color: #0078d4; }
        """)
        edit_btn.clicked.connect(lambda: self.openDeafGeneEditDialog(row))
        return edit_btn
    
    def _buildDeafGeneDeleteBtn(self, sample_id):
        delete_btn = QPushButton("删除")
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #ffffff;
                color: #333333;
                border: 1px solid #f44336;
                border-radius: 3px;
                padding: 4px 8px;
                font-size: 11px;
            }
            QPushButton:hover { background-color: #ffebee; border-color: #d32f2f; }
        """)
        delete_btn.clicked.connect(lambda: self.removeDeafGeneSample(sample_id))
        return delete_btn
    
    def openDeafGeneAddDialog(self):
        dialog = DeafGeneSampleDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            sample_data = dialog.getDeafGeneSampleData()
            sample_data['created_by'] = auth_manager.current_user['id']
            
            sample_no = sample_data['sample_no']
            if not sample_no:
                QMessageBox.warning(self, "输入错误", "样本编号不能为空")
                return
            
            if not sample_data['patient_name']:
                QMessageBox.warning(self, "输入错误", "患者姓名不能为空")
                return
            
            cursor = db.execute_query("SELECT id FROM samples WHERE sample_no = ?", (sample_no,))
            if cursor.fetchone():
                QMessageBox.warning(self, "重复错误", f"样本编号 {sample_no} 已存在")
                return
            
            try:
                db.create_sample(sample_data)
                QMessageBox.information(self, "成功", "样本添加成功")
                self.refreshDeafGeneSampleList()
            except Exception as e:
                QMessageBox.critical(self, "数据库错误", f"添加样本失败: {str(e)}")
    
    def openDeafGeneEditDialog(self, row):
        sample_no = self.deaf_gene_sample_table.item(row, 0).text()
        
        cursor = db.execute_query(
            "SELECT * FROM samples WHERE sample_no = ?",
            (sample_no,)
        )
        sample_row = cursor.fetchone()
        
        if not sample_row:
            QMessageBox.warning(self, "数据错误", "未找到样本信息")
            return
        
        sample = dict(sample_row)
        dialog = DeafGeneSampleDialog(self, sample)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            sample_data = dialog.getDeafGeneSampleData()
            
            patient_name = sample_data['patient_name']
            if not patient_name:
                QMessageBox.warning(self, "输入错误", "患者姓名不能为空")
                return
            
            try:
                db.execute_query("""
                    UPDATE samples SET
                        patient_name = ?, gender = ?, age = ?,
                        clinical_diagnosis = ?, hospital = ?,
                        test_project = ?, family_history = ?, notes = ?,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (
                    patient_name,
                    sample_data.get('gender'),
                    sample_data.get('age'),
                    sample_data.get('clinical_diagnosis'),
                    sample_data.get('hospital'),
                    sample_data.get('test_project'),
                    sample_data.get('family_history'),
                    sample_data.get('notes'),
                    sample['id']
                ))
                db.commit()
                
                QMessageBox.information(self, "成功", "样本信息更新成功")
                self.refreshDeafGeneSampleList()
            except Exception as e:
                QMessageBox.critical(self, "数据库错误", f"更新样本失败: {str(e)}")
    
    def removeDeafGeneSample(self, sample_id):
        reply = QMessageBox.question(
            self, '确认删除',
            '确定要删除这个样本吗？此操作不可恢复。',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                db.execute_query("DELETE FROM gene_data WHERE sample_id = ?", (sample_id,))
                db.execute_query("DELETE FROM reports WHERE sample_id = ?", (sample_id,))
                db.execute_query("DELETE FROM samples WHERE id = ?", (sample_id,))
                db.commit()
                
                QMessageBox.information(self, "成功", "样本删除成功")
                self.refreshDeafGeneSampleList()
            except Exception as e:
                QMessageBox.critical(self, "数据库错误", f"删除样本失败: {str(e)}")
    
    def runDeafGeneSearch(self):
        filters = {}
        
        sample_no = self.deaf_gene_sample_no_input.text().strip()
        if sample_no:
            filters['sample_no'] = sample_no
        
        name = self.deaf_gene_patient_name_input.text().strip()
        if name:
            filters['patient_name'] = name
        
        status = self.deaf_gene_status_filter.currentData()
        if status:
            filters['status'] = status
        
        self._deafGeneCurFilter = filters
        self._deafGeneLastSearchTime = __import__('time').time()
        self.refreshDeafGeneSampleList(filters)
    
    def clearDeafGeneSearchInputs(self):
        self.deaf_gene_sample_no_input.clear()
        self.deaf_gene_patient_name_input.clear()
        self.deaf_gene_status_filter.setCurrentIndex(0)
        self._deafGeneCurFilter = {}
        self.refreshDeafGeneSampleList()
    
    def importDeafGeneExcelData(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        
        if file_path:
            QMessageBox.information(self, "提示", "批量导入功能开发中...")
    
    def exportDeafGeneToExcel(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存样本列表", "", "Excel文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "样本列表"
            
            headers = [
                "样本编号", "受检者姓名", "性别", "年龄", 
                "临床诊断", "送检单位", "检测项目", "状态",
                "创建时间", "创建人"
            ]
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            samples = db.get_samples()
            
            for row, sample in enumerate(samples, 2):
                ws.cell(row=row, column=1, value=sample.get('sample_no', '')).border = thin_border
                ws.cell(row=row, column=2, value=sample.get('patient_name', '')).border = thin_border
                ws.cell(row=row, column=3, value=sample.get('gender', '')).border = thin_border
                ws.cell(row=row, column=4, value=sample.get('age', '')).border = thin_border
                ws.cell(row=row, column=5, value=sample.get('clinical_diagnosis', '')).border = thin_border
                ws.cell(row=row, column=6, value=sample.get('hospital', '')).border = thin_border
                ws.cell(row=row, column=7, value=sample.get('test_project', '')).border = thin_border
                ws.cell(row=row, column=8, value=sample.get('status', '')).border = thin_border
                ws.cell(row=row, column=9, value=sample.get('created_at', '')).border = thin_border
                
                created_by = sample.get('created_by', '')
                if created_by:
                    cursor = db.execute_query("SELECT real_name FROM users WHERE id = ?", (created_by,))
                    user_row = cursor.fetchone()
                    if user_row:
                        created_by_name = dict(user_row).get('real_name', str(created_by))
                    else:
                        created_by_name = str(created_by)
                else:
                    created_by_name = ''
                ws.cell(row=row, column=10, value=created_by_name).border = thin_border
            
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 20
            
            wb.save(file_path)
            QMessageBox.information(self, "成功", f"样本列表已导出到\n{file_path}")
            
        except ImportError:
            QMessageBox.critical(self, "依赖错误", "需要安装 openpyxl 库才能导出Excel")
        except Exception as e:
            QMessageBox.critical(self, "导出错误", f"导出失败: {str(e)}")
    
    def refreshDeafGeneData(self):
        self.refreshDeafGeneSampleList()


class DeafGeneSampleDialog(QDialog):
    def __init__(self, parent=None, sample=None):
        super().__init__(parent)
        self._deafGeneSampleData = sample
        self.initDeafGeneDialogUI()
        
    def initDeafGeneDialogUI(self):
        if self._deafGeneSampleData:
            self.setWindowTitle("编辑样本信息")
        else:
            self.setWindowTitle("新增样本信息")
        
        self.setMinimumWidth(500)
        
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        
        self.deaf_gene_sample_no_input = QLineEdit()
        if self._deafGeneSampleData:
            self.deaf_gene_sample_no_input.setText(self._deafGeneSampleData['sample_no'])
            self.deaf_gene_sample_no_input.setEnabled(False)
        else:
            self.deaf_gene_sample_no_input.setPlaceholderText("请输入样本编号")
        form_layout.addRow("样本编号*:", self.deaf_gene_sample_no_input)
        
        self.deaf_gene_patient_name_input = QLineEdit()
        if self._deafGeneSampleData:
            self.deaf_gene_patient_name_input.setText(self._deafGeneSampleData['patient_name'])
        else:
            self.deaf_gene_patient_name_input.setPlaceholderText("请输入受检者姓名")
        form_layout.addRow("受检者姓名*:", self.deaf_gene_patient_name_input)
        
        self.deaf_gene_gender_combo = QComboBox()
        self.deaf_gene_gender_combo.addItems(["男", "女"])
        if self._deafGeneSampleData:
            self.deaf_gene_gender_combo.setCurrentText(self._deafGeneSampleData.get('gender', '男'))
        form_layout.addRow("性别:", self.deaf_gene_gender_combo)
        
        self.deaf_gene_age_input = QLineEdit()
        self.deaf_gene_age_input.setPlaceholderText("请输入年龄")
        if self._deafGeneSampleData:
            self.deaf_gene_age_input.setText(self._deafGeneSampleData.get('age', ''))
        form_layout.addRow("年龄:", self.deaf_gene_age_input)
        
        self.deaf_gene_diagnosis_input = QLineEdit()
        self.deaf_gene_diagnosis_input.setPlaceholderText("请输入临床诊断")
        if self._deafGeneSampleData:
            self.deaf_gene_diagnosis_input.setText(self._deafGeneSampleData.get('clinical_diagnosis', ''))
        form_layout.addRow("临床诊断:", self.deaf_gene_diagnosis_input)
        
        self.deaf_gene_hospital_input = QLineEdit()
        self.deaf_gene_hospital_input.setPlaceholderText("请输入送检单位")
        if self._deafGeneSampleData:
            self.deaf_gene_hospital_input.setText(self._deafGeneSampleData.get('hospital', ''))
        form_layout.addRow("送检单位:", self.deaf_gene_hospital_input)
        
        self.deaf_gene_project_input = QLineEdit()
        self.deaf_gene_project_input.setPlaceholderText("请输入检测项目")
        if self._deafGeneSampleData:
            self.deaf_gene_project_input.setText(self._deafGeneSampleData.get('test_project', ''))
        form_layout.addRow("检测项目:", self.deaf_gene_project_input)
        
        layout.addLayout(form_layout)
        
        extended_group = QGroupBox("扩展信息")
        extended_layout = QFormLayout()
        
        self.deaf_gene_family_history_input = QTextEdit()
        self.deaf_gene_family_history_input.setMaximumHeight(80)
        self.deaf_gene_family_history_input.setPlaceholderText("请输入家族耳聋史")
        if self._deafGeneSampleData:
            self.deaf_gene_family_history_input.setText(self._deafGeneSampleData.get('family_history', ''))
        extended_layout.addRow("家族耳聋史:", self.deaf_gene_family_history_input)
        
        self.deaf_gene_notes_input = QTextEdit()
        self.deaf_gene_notes_input.setMaximumHeight(80)
        self.deaf_gene_notes_input.setPlaceholderText("请输入备注信息")
        if self._deafGeneSampleData:
            self.deaf_gene_notes_input.setText(self._deafGeneSampleData.get('notes', ''))
        extended_layout.addRow("备注:", self.deaf_gene_notes_input)
        
        extended_group.setLayout(extended_layout)
        layout.addWidget(extended_group)
        
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.validateDeafGeneAndSave)
        button_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        layout.addLayout(button_layout)
    
    def validateDeafGeneAndSave(self):
        sample_no = self.deaf_gene_sample_no_input.text().strip()
        if not sample_no:
            QMessageBox.warning(self, "验证失败", "请输入样本编号")
            return
        
        patient_name = self.deaf_gene_patient_name_input.text().strip()
        if not patient_name:
            QMessageBox.warning(self, "验证失败", "请输入受检者姓名")
            return
        
        self.accept()
    
    def getDeafGeneSampleData(self):
        return {
            'sample_no': self.deaf_gene_sample_no_input.text().strip(),
            'patient_name': self.deaf_gene_patient_name_input.text().strip(),
            'gender': self.deaf_gene_gender_combo.currentText(),
            'age': self.deaf_gene_age_input.text().strip(),
            'clinical_diagnosis': self.deaf_gene_diagnosis_input.text().strip(),
            'hospital': self.deaf_gene_hospital_input.text().strip(),
            'test_project': self.deaf_gene_project_input.text().strip(),
            'family_history': self.deaf_gene_family_history_input.toPlainText().strip(),
            'notes': self.deaf_gene_notes_input.toPlainText().strip()
        }